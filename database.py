from __future__ import annotations

import sqlite3
import threading
from datetime import date, datetime
from pathlib import Path
from typing import Iterable

from app_config import DB_PATH, EXCEL_ORIGEM, SQLITE_BUSY_TIMEOUT_MS
from app_logging import get_logger

try:
    from openpyxl import load_workbook
except ModuleNotFoundError:
    load_workbook = None


LOGGER = get_logger(__name__)

ABAS_EXCEL = {
    "motoristas": "nome",
    "fazendas": "fazendas",
    "variedades": "variedades",
}

NOTA_FIELDS = (
    "numero",
    "motorista_cod",
    "motorista_nome",
    "caminhao",
    "operador_cod",
    "operador_nome",
    "colhedora",
    "faz_muda_cod",
    "faz_muda_nome",
    "talhao",
    "faz_plantio_cod",
    "faz_plantio_nome",
    "variedade_id",
    "variedade_nome",
    "data_colheita",
    "data_plantio",
)

REFERENCE_TABLES = {
    "motoristas": {"id_col": "codigo", "label": "motorista"},
    "fazendas": {"id_col": "codigo", "label": "fazenda"},
    "variedades": {"id_col": "id", "label": "variedade"},
}

CORRUPTION_MARKERS = (
    "database disk image is malformed",
    "file is not a database",
    "database schema is malformed",
)


class DatabaseCorruptionError(sqlite3.DatabaseError):
    def __init__(self, path: str | Path, details: str):
        self.path = Path(path)
        self.details = str(details).strip() or "Falha desconhecida na integridade do banco."
        super().__init__(f"Banco corrompido em '{self.path}': {self.details}")


def _is_corruption_message(message: str) -> bool:
    message_lc = str(message).lower()
    return any(marker in message_lc for marker in CORRUPTION_MARKERS)


def _sqlite_sidecar_paths(db_path: str | Path) -> tuple[Path, ...]:
    path = Path(db_path)
    return (
        path.with_name(f"{path.name}-wal"),
        path.with_name(f"{path.name}-shm"),
        path.with_name(f"{path.name}-journal"),
    )


def sqlite_integrity_status(db_path: str | Path) -> str:
    path = Path(db_path)
    if not path.exists():
        raise FileNotFoundError(f"Banco de dados nao encontrado: {path}")

    conn = sqlite3.connect(path)
    try:
        row = conn.execute("PRAGMA quick_check;").fetchone()
    except sqlite3.DatabaseError as exc:
        return str(exc)
    finally:
        conn.close()

    if not row or row[0] is None:
        return "quick_check sem retorno"
    return str(row[0]).strip()


def ensure_sqlite_integrity(db_path: str | Path) -> None:
    status = sqlite_integrity_status(db_path)
    if status.lower() != "ok":
        raise DatabaseCorruptionError(db_path, status)


def _quarantine_sqlite_files(db_path: str | Path) -> dict[Path, Path]:
    path = Path(db_path)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    moved: dict[Path, Path] = {}

    for candidate in (path, *_sqlite_sidecar_paths(path)):
        if not candidate.exists():
            continue

        quarantined = candidate.with_name(f"{candidate.name}.corrompido_{timestamp}")
        sequence = 1
        while quarantined.exists():
            quarantined = candidate.with_name(
                f"{candidate.name}.corrompido_{timestamp}_{sequence}"
            )
            sequence += 1

        candidate.replace(quarantined)
        moved[candidate] = quarantined

    return moved


def _restore_quarantined_files(moved: dict[Path, Path]) -> None:
    for original, quarantined in reversed(list(moved.items())):
        if original.exists() or not quarantined.exists():
            continue
        quarantined.replace(original)


def _normalize_note_date(value, field_label: str) -> str | None:
    if value in (None, ""):
        return value

    if isinstance(value, datetime):
        parsed = value.date()
    elif isinstance(value, date):
        parsed = value
    else:
        try:
            parsed = date.fromisoformat(str(value).strip())
        except ValueError as exc:
            raise ValueError(f"{field_label} invalida: use o formato AAAA-MM-DD.") from exc

    today = date.today()
    if parsed > today:
        today_br = today.strftime("%d/%m/%Y")
        raise ValueError(
            f"{field_label} nao pode ser maior que a data atual do sistema ({today_br})."
        )

    return parsed.isoformat()


def restore_sqlite_backup(
    destination_path: str | Path,
    backup_path: str | Path,
) -> tuple[Path, Path | None]:
    destination = Path(destination_path)
    origem = Path(backup_path)
    if not origem.exists():
        raise FileNotFoundError(f"Backup nao encontrado: {origem}")

    ensure_sqlite_integrity(origem)
    destination.parent.mkdir(parents=True, exist_ok=True)

    moved = _quarantine_sqlite_files(destination)
    quarantined_main = moved.get(destination)

    try:
        backup_sqlite_file(origem, destination)
        ensure_sqlite_integrity(destination)
    except Exception:
        try:
            if destination.exists():
                destination.unlink()
        finally:
            _restore_quarantined_files(moved)
        raise

    return destination, quarantined_main


def backup_sqlite_file(source_path: str | Path, destination_path: str | Path) -> Path:
    source = Path(source_path)
    destination = Path(destination_path)

    if not source.exists():
        raise FileNotFoundError(f"Banco de dados nao encontrado: {source}")

    destination.parent.mkdir(parents=True, exist_ok=True)

    src = sqlite3.connect(source)
    dst = sqlite3.connect(destination)
    try:
        src.backup(dst)
        dst.commit()
    finally:
        dst.close()
        src.close()

    return destination


class DB:
    def __init__(self, path: str | Path = DB_PATH, seed_from_excel: bool = True):
        self.path = Path(path)
        self.path.parent.mkdir(parents=True, exist_ok=True)
        self._lock = threading.RLock()
        need_seed = seed_from_excel and (
            not self.path.exists() or self.path.stat().st_size == 0
        )

        self.conn = self._open_connection()
        try:
            self._validate_integrity()
            self._criar_tabelas()
            self._criar_indices()

            if need_seed:
                try:
                    self._seed_from_excel()
                except Exception:
                    LOGGER.exception("Falha ao importar dados iniciais do Excel")
        except Exception:
            self.close()
            raise

    def _open_connection(self) -> sqlite3.Connection:
        conn = sqlite3.connect(
            self.path,
            check_same_thread=False,
            timeout=SQLITE_BUSY_TIMEOUT_MS / 1000,
        )
        conn.row_factory = sqlite3.Row
        try:
            self._configure_connection(conn)
        except sqlite3.DatabaseError as exc:
            conn.close()
            raise self._wrap_database_error(exc) from exc
        return conn

    def _validate_integrity(self) -> None:
        with self._lock:
            try:
                row = self.conn.execute("PRAGMA quick_check;").fetchone()
            except sqlite3.DatabaseError as exc:
                raise DatabaseCorruptionError(self.path, str(exc)) from exc

        status = str(row[0]).strip() if row and row[0] is not None else "quick_check sem retorno"
        if status.lower() != "ok":
            raise DatabaseCorruptionError(self.path, status)

    def _wrap_database_error(self, exc: sqlite3.DatabaseError) -> sqlite3.DatabaseError:
        if isinstance(exc, DatabaseCorruptionError):
            return exc
        if _is_corruption_message(str(exc)):
            return DatabaseCorruptionError(self.path, str(exc))
        return exc

    def _configure_connection(self, conn: sqlite3.Connection) -> None:
        conn.execute("PRAGMA foreign_keys = ON;")
        conn.execute(f"PRAGMA busy_timeout = {SQLITE_BUSY_TIMEOUT_MS};")
        conn.execute("PRAGMA journal_mode = WAL;")
        conn.execute("PRAGMA synchronous = NORMAL;")

    def _fetchone(self, sql: str, params: Iterable = ()) -> sqlite3.Row | None:
        with self._lock:
            try:
                return self.conn.execute(sql, tuple(params)).fetchone()
            except sqlite3.DatabaseError as exc:
                raise self._wrap_database_error(exc) from exc

    def _fetchall(self, sql: str, params: Iterable = ()) -> list[sqlite3.Row]:
        with self._lock:
            try:
                return self.conn.execute(sql, tuple(params)).fetchall()
            except sqlite3.DatabaseError as exc:
                raise self._wrap_database_error(exc) from exc

    def _execute(self, sql: str, params: Iterable = (), commit: bool = False) -> sqlite3.Cursor:
        with self._lock:
            try:
                cursor = self.conn.execute(sql, tuple(params))
                if commit:
                    self.conn.commit()
                return cursor
            except sqlite3.DatabaseError as exc:
                raise self._wrap_database_error(exc) from exc

    def _executemany(self, sql: str, params: Iterable[Iterable], commit: bool = False) -> sqlite3.Cursor:
        with self._lock:
            try:
                cursor = self.conn.executemany(sql, params)
                if commit:
                    self.conn.commit()
                return cursor
            except sqlite3.DatabaseError as exc:
                raise self._wrap_database_error(exc) from exc

    def _table_config(self, tabela: str) -> dict[str, str]:
        if tabela not in REFERENCE_TABLES:
            raise ValueError(f"Tabela de referencia invalida: {tabela}")
        return REFERENCE_TABLES[tabela]

    def _criar_tabelas(self) -> None:
        with self._lock:
            cur = self.conn.cursor()
            cur.executescript(
                """
                CREATE TABLE IF NOT EXISTS motoristas (
                    codigo INTEGER PRIMARY KEY,
                    nome   TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS fazendas (
                    codigo TEXT PRIMARY KEY,
                    nome   TEXT NOT NULL,
                    lat    REAL,
                    lon    REAL
                );
                CREATE TABLE IF NOT EXISTS variedades (
                    id     INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome   TEXT NOT NULL UNIQUE
                );
                CREATE TABLE IF NOT EXISTS notas (
                    numero           INTEGER PRIMARY KEY,
                    motorista_cod    INTEGER,
                    motorista_nome   TEXT,
                    caminhao         TEXT,
                    operador_cod     INTEGER,
                    operador_nome    TEXT,
                    colhedora        TEXT,
                    faz_muda_cod     TEXT,
                    faz_muda_nome    TEXT,
                    talhao           TEXT,
                    faz_plantio_cod  TEXT,
                    faz_plantio_nome TEXT,
                    variedade_id     INTEGER,
                    variedade_nome   TEXT,
                    data_colheita    TEXT,
                    data_plantio     TEXT,
                    duplicado        INTEGER DEFAULT 0
                );

                CREATE TABLE IF NOT EXISTS cidades (
                    id   INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL,
                    lat  REAL,
                    lon  REAL
                );

                CREATE TABLE IF NOT EXISTS frota (
                    id     INTEGER PRIMARY KEY AUTOINCREMENT,
                    numero TEXT NOT NULL UNIQUE,
                    placa  TEXT,
                    status TEXT DEFAULT 'ATIVO'
                );

                CREATE TABLE IF NOT EXISTS frentes (
                    id   INTEGER PRIMARY KEY AUTOINCREMENT,
                    nome TEXT NOT NULL UNIQUE
                );

                CREATE TABLE IF NOT EXISTS localizacao_frentes (
                    id          INTEGER PRIMARY KEY AUTOINCREMENT,
                    data        TEXT NOT NULL,
                    frente_id   INTEGER NOT NULL,
                    fazenda_cod TEXT NOT NULL,
                    FOREIGN KEY(frente_id) REFERENCES frentes(id),
                    FOREIGN KEY(fazenda_cod) REFERENCES fazendas(codigo)
                );

                CREATE TABLE IF NOT EXISTS escala_viagem (
                    id         INTEGER PRIMARY KEY AUTOINCREMENT,
                    onibus_id  INTEGER NOT NULL,
                    frente_id  INTEGER NOT NULL,
                    cidade_id  INTEGER NOT NULL,
                    hora_ida   TEXT NOT NULL,
                    hora_volta TEXT NOT NULL,
                    FOREIGN KEY(onibus_id) REFERENCES frota(id),
                    FOREIGN KEY(frente_id) REFERENCES frentes(id),
                    FOREIGN KEY(cidade_id) REFERENCES cidades(id)
                );
                """
            )
            self.conn.commit()

    def _criar_indices(self) -> None:
        with self._lock:
            self.conn.executescript(
                """
                CREATE INDEX IF NOT EXISTS idx_notas_data_colheita ON notas(data_colheita);
                CREATE INDEX IF NOT EXISTS idx_notas_motorista_nome ON notas(motorista_nome);
                CREATE INDEX IF NOT EXISTS idx_notas_operador_nome ON notas(operador_nome);
                CREATE INDEX IF NOT EXISTS idx_notas_faz_plantio_nome ON notas(faz_plantio_nome);
                CREATE INDEX IF NOT EXISTS idx_notas_faz_muda_nome ON notas(faz_muda_nome);
                CREATE INDEX IF NOT EXISTS idx_notas_variedade_nome ON notas(variedade_nome);
                CREATE INDEX IF NOT EXISTS idx_fazendas_nome ON fazendas(nome);
                CREATE INDEX IF NOT EXISTS idx_motoristas_nome ON motoristas(nome);
                CREATE INDEX IF NOT EXISTS idx_variedades_nome ON variedades(nome);
                CREATE INDEX IF NOT EXISTS idx_localizacao_frentes_data ON localizacao_frentes(data);
                """
            )
            self.conn.commit()

    def _seed_from_excel(self) -> None:
        if not EXCEL_ORIGEM.exists():
            LOGGER.info("Excel de origem nao encontrado em %s; seed inicial ignorado", EXCEL_ORIGEM)
            return
        if load_workbook is None:
            LOGGER.warning("openpyxl indisponivel; seed inicial do Excel foi ignorado")
            return

        wb = load_workbook(EXCEL_ORIGEM, data_only=True, keep_vba=True)
        try:
            if ABAS_EXCEL["motoristas"] in wb.sheetnames:
                ws = wb[ABAS_EXCEL["motoristas"]]
                lista = []
                for row in ws.iter_rows(min_row=2, max_col=2):
                    try:
                        if row[0].value and row[1].value:
                            lista.append((int(row[0].value), str(row[1].value).strip()))
                    except (TypeError, ValueError):
                        continue
                if lista:
                    self._executemany(
                        "INSERT OR IGNORE INTO motoristas(codigo, nome) VALUES (?, ?)",
                        lista,
                        commit=True,
                    )

            if ABAS_EXCEL["fazendas"] in wb.sheetnames:
                ws = wb[ABAS_EXCEL["fazendas"]]
                lista = []
                for row in ws.iter_rows(min_row=2, max_col=2):
                    if row[0].value and row[1].value:
                        lista.append((str(row[0].value).strip(), str(row[1].value).strip()))
                if lista:
                    self._executemany(
                        "INSERT OR IGNORE INTO fazendas(codigo, nome) VALUES (?, ?)",
                        lista,
                        commit=True,
                    )

            if ABAS_EXCEL["variedades"] in wb.sheetnames:
                ws = wb[ABAS_EXCEL["variedades"]]
                lista = [(str(cell.value).strip(),) for cell in ws["A"][1:] if cell.value]
                if lista:
                    self._executemany(
                        "INSERT OR IGNORE INTO variedades(nome) VALUES (?)",
                        lista,
                        commit=True,
                    )
        finally:
            wb.close()

    def listar_referencias(self, tabela: str) -> list[sqlite3.Row]:
        config = self._table_config(tabela)
        return self._fetchall(
            f"SELECT {config['id_col']}, nome FROM {tabela} ORDER BY nome"
        )

    def listar_motoristas(self) -> list[sqlite3.Row]:
        return self.listar_referencias("motoristas")

    def listar_fazendas(self) -> list[sqlite3.Row]:
        return self.listar_referencias("fazendas")

    def listar_variedades(self) -> list[sqlite3.Row]:
        return self.listar_referencias("variedades")

    def listar_fazendas_com_gps(self, limit: int | None = None, randomize: bool = False) -> list[sqlite3.Row]:
        sql = """
            SELECT nome, lat, lon
            FROM fazendas
            WHERE lat IS NOT NULL
              AND lon IS NOT NULL
        """
        if randomize:
            sql += " ORDER BY RANDOM()"
        else:
            sql += " ORDER BY nome"
        params: tuple[int, ...] = ()
        if limit:
            sql += " LIMIT ?"
            params = (int(limit),)
        return self._fetchall(sql, params)

    def buscar_por_codigo(self, tabela: str, codigo) -> dict | None:
        config = self._table_config(tabela)
        if codigo in (None, ""):
            return None

        query = f"SELECT {config['id_col']} AS codigo, nome FROM {tabela} WHERE {config['id_col']} = ?"
        row = self._fetchone(query, (codigo,))
        if row:
            return dict(row)

        if tabela == "fazendas" and "-" not in str(codigo) and len(str(codigo)) >= 4:
            codigo_txt = str(codigo)
            codigo_formatado = codigo_txt[:3] + "-" + codigo_txt[3:]
            row = self._fetchone(query, (codigo_formatado,))
            if row:
                return dict(row)
        return None

    def buscar_referencia(self, tabela: str, valor, col_id: str | None = None) -> sqlite3.Row | None:
        if valor in (None, ""):
            return None

        config = self._table_config(tabela)
        campo_id = col_id or config["id_col"]
        valor_txt = str(valor).strip()

        row = self._fetchone(
            f"SELECT * FROM {tabela} WHERE {campo_id} = ?",
            (valor_txt,),
        )
        if row:
            return row

        if tabela == "fazendas" and campo_id == "codigo" and "-" not in valor_txt and len(valor_txt) >= 4:
            codigo_formatado = valor_txt[:3] + "-" + valor_txt[3:]
            row = self._fetchone(
                f"SELECT * FROM {tabela} WHERE {campo_id} = ?",
                (codigo_formatado,),
            )
            if row:
                return row

        return self._fetchone(
            f"SELECT * FROM {tabela} WHERE UPPER(TRIM(nome)) = UPPER(TRIM(?))",
            (valor_txt,),
        )

    def resolver_referencia(self, tabela: str, texto: str, col_id: str | None = None) -> sqlite3.Row | None:
        valor = (texto or "").strip()
        if not valor:
            return None

        if " - " in valor:
            codigo, nome = valor.split(" - ", 1)
            return self.buscar_referencia(tabela, codigo.strip(), col_id=col_id) or self.buscar_referencia(
                tabela,
                nome.strip(),
                col_id=col_id,
            )

        return self.buscar_referencia(tabela, valor, col_id=col_id)

    def buscar_nota(self, numero) -> sqlite3.Row | None:
        return self._fetchone("SELECT * FROM notas WHERE numero = ?", (numero,))

    def montar_nota_payload(self, dados: dict) -> dict:
        payload = {campo: dados.get(campo) for campo in NOTA_FIELDS}
        payload["numero"] = int(payload["numero"])
        payload["data_colheita"] = _normalize_note_date(
            payload.get("data_colheita"),
            "Data de colheita",
        )
        payload["data_plantio"] = _normalize_note_date(
            payload.get("data_plantio"),
            "Data de plantio",
        )
        payload["duplicado"] = int(bool(dados.get("duplicado", 0)))
        return payload

    def inserir_nota(self, dados: dict, force: bool = False) -> None:
        payload = self.montar_nota_payload(dados)
        params = [payload[campo] for campo in NOTA_FIELDS] + [payload["duplicado"]]
        columns = ", ".join((*NOTA_FIELDS, "duplicado"))
        placeholders = ", ".join("?" for _ in range(len(NOTA_FIELDS) + 1))

        with self._lock:
            try:
                self.conn.execute(
                    f"INSERT INTO notas ({columns}) VALUES ({placeholders})",
                    params,
                )
            except sqlite3.IntegrityError:
                if not force:
                    raise

                update_values = params[1:] + [payload["numero"]]
                self.conn.execute(
                    """
                    UPDATE notas
                    SET motorista_cod=?, motorista_nome=?, caminhao=?, operador_cod=?, operador_nome=?,
                        colhedora=?, faz_muda_cod=?, faz_muda_nome=?, talhao=?,
                        faz_plantio_cod=?, faz_plantio_nome=?, variedade_id=?, variedade_nome=?,
                        data_colheita=?, data_plantio=?, duplicado=?
                    WHERE numero=?
                    """,
                    update_values,
                )
            self.conn.commit()

    def listar_notas(self, data_inicio: str | None = None, data_fim: str | None = None) -> list[sqlite3.Row]:
        sql = """
            SELECT numero, motorista_cod, motorista_nome, caminhao, operador_cod,
                   operador_nome, colhedora, faz_muda_cod, faz_muda_nome, talhao,
                   faz_plantio_cod, faz_plantio_nome, variedade_nome, data_colheita,
                   data_plantio
            FROM notas
        """
        params: tuple[str, str] | tuple[()] = ()
        if data_inicio and data_fim:
            sql += " WHERE data_colheita BETWEEN ? AND ?"
            params = (data_inicio, data_fim)
        sql += " ORDER BY numero DESC"
        return self._fetchall(sql, params)

    def excluir_nota(self, numero) -> None:
        self._execute("DELETE FROM notas WHERE numero = ?", (numero,), commit=True)

    def contar_notas_total(self) -> int:
        row = self._fetchone("SELECT COUNT(*) AS total FROM notas")
        return int(row["total"] if row else 0)

    def contar_notas_data(self, data_sql: str) -> int:
        row = self._fetchone(
            "SELECT COUNT(*) AS total FROM notas WHERE data_colheita = ?",
            (data_sql,),
        )
        return int(row["total"] if row else 0)

    def contar_notas_periodo(self, d_ini: str, d_fim: str) -> int:
        row = self._fetchone(
            "SELECT COUNT(*) AS total FROM notas WHERE data_colheita BETWEEN ? AND ?",
            (d_ini, d_fim),
        )
        return int(row["total"] if row else 0)

    def contar_dias_ativos_periodo(self, d_ini: str, d_fim: str) -> int:
        row = self._fetchone(
            """
            SELECT COUNT(DISTINCT data_colheita) AS total
            FROM notas
            WHERE data_colheita BETWEEN ? AND ?
            """,
            (d_ini, d_fim),
        )
        return int(row["total"] if row else 0)

    def top_por_coluna(self, coluna: str, d_ini: str, d_fim: str, limit: int = 5) -> list[sqlite3.Row]:
        allowed_cols = {
            "motorista_nome",
            "operador_nome",
            "colhedora",
            "variedade_nome",
            "faz_muda_nome",
            "faz_plantio_nome",
        }
        if coluna not in allowed_cols:
            raise ValueError(f"Coluna nao permitida para ranking: {coluna}")

        return self._fetchall(
            f"""
            SELECT {coluna} AS nome, COUNT(*) AS qtd
            FROM notas
            WHERE data_colheita BETWEEN ? AND ?
              AND {coluna} IS NOT NULL
              AND TRIM({coluna}) <> ''
            GROUP BY {coluna}
            ORDER BY qtd DESC, {coluna}
            LIMIT ?
            """,
            (d_ini, d_fim, int(limit)),
        )

    def top_motorista_do_dia(self, data_sql: str) -> sqlite3.Row | None:
        return self._fetchone(
            """
            SELECT motorista_nome, COUNT(*) AS qtd
            FROM notas
            WHERE data_colheita = ?
              AND motorista_nome IS NOT NULL
              AND TRIM(motorista_nome) <> ''
            GROUP BY motorista_nome
            ORDER BY qtd DESC, motorista_nome
            LIMIT 1
            """,
            (data_sql,),
        )

    def dataframe_por_query(self, sql: str, params: Iterable = ()):
        import pandas as pd

        conn = sqlite3.connect(self.path)
        try:
            return pd.read_sql_query(sql, conn, params=tuple(params))
        finally:
            conn.close()

    def dataframe_historico(self, data_inicio: str | None = None, data_fim: str | None = None):
        sql = "SELECT * FROM notas"
        params: tuple[str, str] | tuple[()] = ()
        if data_inicio and data_fim:
            sql += " WHERE data_colheita BETWEEN ? AND ?"
            params = (data_inicio, data_fim)
        sql += " ORDER BY numero DESC"
        return self.dataframe_por_query(sql, params)

    def dataframe_fluxo(self, d_ini: str, d_fim: str):
        return self.dataframe_por_query(
            """
            SELECT data_colheita AS Data, faz_muda_nome AS Origem,
                   talhao AS Talhao, variedade_nome AS Variedade,
                   faz_plantio_nome AS Destino, COUNT(*) AS Qtd
            FROM notas
            WHERE data_colheita BETWEEN ? AND ?
            GROUP BY data_colheita, faz_muda_nome, talhao, variedade_nome, faz_plantio_nome
            ORDER BY data_colheita
            """,
            (d_ini, d_fim),
        )

    def cadastrar_novo(self, tipo: str, codigo, nome: str):
        try:
            if tipo == "motorista":
                self.adicionar_motorista(codigo, nome)
            elif tipo == "fazenda":
                self.adicionar_fazenda(codigo, nome)
            elif tipo == "variedade":
                self.adicionar_variedade(nome)
            else:
                raise ValueError(f"Tipo de cadastro invalido: {tipo}")
            return True, "Sucesso"
        except Exception as exc:
            return False, str(exc)

    def adicionar_motorista(self, codigo, nome: str) -> None:
        self._execute(
            "INSERT INTO motoristas (codigo, nome) VALUES (?, ?)",
            (int(codigo), str(nome).strip()),
            commit=True,
        )

    def adicionar_fazenda(self, codigo, nome: str) -> None:
        self._execute(
            "INSERT INTO fazendas (codigo, nome) VALUES (?, ?)",
            (str(codigo).strip(), str(nome).strip()),
            commit=True,
        )

    def adicionar_variedade(self, nome: str) -> None:
        self._execute(
            "INSERT INTO variedades (nome) VALUES (?)",
            (str(nome).strip(),),
            commit=True,
        )

    def excluir_cadastro(self, tabela: str, valor_id) -> None:
        config = self._table_config(tabela)
        self._execute(
            f"DELETE FROM {tabela} WHERE {config['id_col']} = ?",
            (valor_id,),
            commit=True,
        )

    def create_backup(self, destination_path: str | Path) -> Path:
        destination = Path(destination_path)
        destination.parent.mkdir(parents=True, exist_ok=True)

        with self._lock:
            self.conn.execute("PRAGMA wal_checkpoint(PASSIVE);")
            dst = sqlite3.connect(destination)
            try:
                self.conn.backup(dst)
                dst.commit()
            finally:
                dst.close()

        return destination

    def restore_from_backup(self, backup_path: str | Path) -> None:
        origem = Path(backup_path)
        if not origem.exists():
            raise FileNotFoundError(f"Backup nao encontrado: {origem}")

        try:
            with self._lock:
                self.conn.close()
                restore_sqlite_backup(self.path, origem)
                self.conn = self._open_connection()
                self._validate_integrity()
        except sqlite3.DatabaseError as exc:
            raise self._wrap_database_error(exc) from exc

    def resumo_gps_fazendas(self) -> dict[str, int]:
        total = self._fetchone("SELECT COUNT(*) AS total FROM fazendas")
        com_gps = self._fetchone(
            """
            SELECT COUNT(*) AS total
            FROM fazendas
            WHERE lat IS NOT NULL
              AND lon IS NOT NULL
            """
        )
        total_int = int(total["total"] if total else 0)
        com_gps_int = int(com_gps["total"] if com_gps else 0)
        return {
            "total": total_int,
            "com_gps": com_gps_int,
            "sem_gps": max(0, total_int - com_gps_int),
        }

    def close(self) -> None:
        with self._lock:
            conn = getattr(self, "conn", None)
            if conn is not None:
                conn.close()
                self.conn = None
